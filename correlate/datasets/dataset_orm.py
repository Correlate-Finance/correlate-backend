from .models import DatasetMetadata, Dataset
from datetime import datetime, UTC
import openpyxl
from django.core.files.uploadedfile import UploadedFile
import pytz
import pandas as pd
from dateutil.parser import parse
from frozendict import frozendict
from core.data_processing import transform_data_base

CACHED_DFS = None


def add_dataset_bulk(records: list[tuple[datetime, float]], metadata: DatasetMetadata):
    to_add = []

    existing_datasets = Dataset.objects.filter(metadata=metadata).values_list(
        "date", "value"
    )
    existing_ds_map = {ds[0]: ds[1] for ds in existing_datasets}
    for record in records:
        date = record[0]
        date = date.replace(tzinfo=UTC)
        value = record[1]

        if existing_ds_map.get(date, None) == value:
            # Record already exists skip
            continue

        to_add.append(Dataset(metadata=metadata, date=date, value=value))
    added = len(Dataset.objects.bulk_create(to_add, ignore_conflicts=True))
    return added


def parse_excel_file_for_datasets(excel_file: UploadedFile):
    workbook = openpyxl.load_workbook(filename=excel_file, data_only=True)

    results = []

    for sheet in workbook:
        # Step 1: Parse Metadata
        metadata = {}

        dataset: list[tuple[datetime, float]] = []
        data_start = False
        for row in sheet.iter_rows():
            col1 = row[0].value
            col2 = row[1].value

            if (
                isinstance(col1, str)
                and col1.lower() == "date"
                and isinstance(col2, str)
                and col2.lower() == "value"
            ):
                data_start = True  # Found the dataset header
                continue

            if not data_start:
                # Until data starts we keep parsing metadata
                key, value = [
                    str(cell.value) for cell in row[:2]
                ]  # Assuming key-value in first two columns
                if key and value:  # Check if both key and value are present
                    metadata[key.lower()] = value

            else:
                raw_date, value = str(col1), col2

                if raw_date and value:  # Check if both date and value are present
                    date = (
                        parse(raw_date).replace(tzinfo=pytz.utc)
                        if isinstance(raw_date, str)
                        else raw_date
                    )
                    dataset.append((date, float(value)))  # type:ignore

                if not col1 and not col2:
                    # Stop parsing if we find an empty row after the dataset
                    break

        dataset_metadata, created = DatasetMetadata.objects.get_or_create(
            internal_name=sheet.title,
            defaults=dict(
                external_name=metadata.get("name", sheet.title),
                source=metadata.get("source", None),
                description=metadata.get("description", None),
            ),
        )

        total_new = add_dataset_bulk(dataset, dataset_metadata)
        results.append((sheet.title, created, total_new))
    return results


def parse_metadata_from_excel(excel_file: UploadedFile):
    workbook = openpyxl.load_workbook(filename=excel_file, data_only=True)
    total = 0
    results = []
    for sheet in workbook:
        headers = []
        first_row = True
        filter_field = "internal_name"
        for row in sheet.iter_rows():
            if first_row:
                headers = [str(cell.value).lower() for cell in row]
                first_row = False
                filter_field = (
                    "external_name"
                    if headers[0] == "external_name"
                    else "internal_name"
                )
                continue

            updates = {}
            dm = None
            name = ""
            skip_row = False
            for i, cell in enumerate(row):
                if i == 0:
                    if cell.value is None:
                        skip_row = True
                        break

                    dm = DatasetMetadata.objects.filter(
                        **{filter_field: cell.value, "hidden": False}
                    )
                    name = cell.value

                else:
                    if headers[i] == "categories":
                        updates[headers[i]] = [
                            category.strip() for category in str(cell.value).split(",")
                        ]
                    else:
                        updates[headers[i]] = cell.value

            if skip_row:
                continue

            # Make sure that the Dataset Metadata was correctly fetched
            if dm is None:
                results.append((name, "Invalid row"))
            elif dm.count() > 1:
                results.append((name, "Multiple metadata found"))
            elif not dm.exists():
                results.append((name, "Metadata not found"))
            else:
                dm.update(**updates)
                total += 1

    return [("success", f"Updated {total} metadata records")] + results


def get_all_dfs(
    selected_names: list[str] | None = None,
) -> frozendict[str, pd.DataFrame]:
    global CACHED_DFS
    if CACHED_DFS:
        if selected_names is None:
            return CACHED_DFS
        else:
            return frozendict(
                {
                    name: CACHED_DFS[name]
                    for name in selected_names
                    if name in CACHED_DFS
                }
            )

    dfs = {}
    if selected_names is not None:
        datasets = Dataset.objects.filter(
            metadata__internal_name__in=selected_names
        ).prefetch_related("metadata")
    else:
        dataset_metadatas = (
            DatasetMetadata.objects.filter(hidden=False)
            .order_by("created_at")
            .values_list("id", flat=True)
        )
        datasets = Dataset.objects.filter(
            metadata_id__in=dataset_metadatas
        ).prefetch_related("metadata")

    datasets = datasets.values_list("metadata__internal_name", "date", "value")
    for dataset in datasets:
        title = dataset[0]
        if title not in dfs:
            dfs[title] = []
        dfs[title].append((dataset[1], dataset[2]))

    for title, data in dfs.items():
        dfs[title] = pd.DataFrame(data, columns=["Date", "Value"])
        transform_data_base(dfs[title])

    if selected_names is None:
        CACHED_DFS = dfs
    return frozendict(dfs)


def get_df(title: str) -> pd.DataFrame | None:
    if CACHED_DFS:
        return CACHED_DFS.get(title)
    dataset = list(Dataset.objects.filter(metadata__internal_name=title).all())
    if len(dataset) == 0:
        return None
    data = [(d.date, d.value) for d in dataset]
    return pd.DataFrame(data, columns=["Date", "Value"])
